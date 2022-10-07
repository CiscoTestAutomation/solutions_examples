from pyats.topology import Testbed, Device
from genie.testbed import load
from multiprocessing.dummy import Pool as ThreadPool
import datetime
from datetime import datetime
import argparse
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import logging
import os

logging.basicConfig(filename='log_ccollect.log', level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s : %(message)s')


def Testbed_routine(hostname):

    uut = tb.devices[hostname]
    connected=False
    try:
        uut.connect(init_config_commands=[],connection_timeout=5,log_stdout=True)
        connected=True
    except:
        logging.debug("Error #1 in the connection to: '{}'".format(hostname))

    if connected:
        print('Device: "{}" connected, collecting'.format(hostname))
        try:
            output = uut.parse('show ipv4 interface brief')
            ip_underlay = output['interface']['Loopback0']['ip_address']
            sheet.append([hostname,ip_underlay])
        except Exception as e:
            logging.debug('Exception catched on Hostname: "{}"\n{}'.format(hostname,e))


        print('Processing complete for device: "{}"'.format(hostname))


def pool_connection(maxthreads,hostname):
    start_time = datetime.strftime(datetime.now(), "%d/%m/%y %H:%M:%S")
    pool = ThreadPool(maxthreads)
    pool_result = []
    try:
        #pool.map (funcion to iterate)
        pool_result = pool.map(Testbed_routine,hostname)
    except UnboundLocalError:
        print("Error mapping threads to routine")

    pool.close()
    pool.join()

    end_time = datetime.strftime(datetime.now(), "%d/%m/%y %H:%M:%S")
    print("Info - Start Time:", start_time)
    print("Info - End Time:", end_time)
    print("")

    return pool_result

def Create_Testbed(user,pswd,hostnames,device_ips,jump_bool,jump_ip):
    testbed = Testbed('exampleTestbed')

    if jump_bool:
        jump_server = Device('jump_host',
                    connections = {
                        'mgmt': {
                            'protocol': 'ssh',
                            'ip': jump_ip,
                            'port': 10104
                        },
                    })
        jump_server.os = 'linux'
        jump_server.testbed = testbed
        jump_server.credentials['default'] = dict(username='cisco', password='pyats2022')

    for i in range(len(hostnames)):
        if jump_bool:
            dev = Device(hostnames[i],
                connections = {
                            'cli': {
                                'protocol': 'ssh',
                                'ip': device_ips[i],
                                'proxy': 'jump_host',
                                'login_creds': ['default', 'local']
                                },
                            })
        else:
            dev = Device(hostnames[i],
            connections = {
                        'cli': {
                            'protocol': 'ssh',
                            'ip': device_ips[i],
                            'login_creds': ['default']
                            },
                        })
        dev.os = 'iosxr'
        dev.testbed = testbed
        dev.credentials['default'] = dict(username=user, password=pswd)
        logging.debug('Device {} added to testbed'.format(hostnames[i]))
        del dev
    return testbed

def main():
    parser=argparse.ArgumentParser()
    parser.add_argument('--file', help='Input File', required=True)
    parser.add_argument('--credentials', help='Send vRouter creds in format user:pass', required=True)
    parser.add_argument('--jump', help='Send Jumphost IP', required=False)
    args=parser.parse_args()

    file = args.file
    creds = args.credentials
    jump_bool = False
    jump_ip = ""
    if args.jump :
        jump_bool = True
        jump_ip = args.jump

    user = creds.split(":")[0]
    password = creds.split(":")[1]

    global full_table
    full_table = load_workbook(file)
    devices = full_table['Device Info']

    device_ips = []
    hostnames = []

    for i in range(1, len(list(list(devices.columns)[0]))):
        hostnames.append(list(list(devices.columns)[0])[i].value)

    for i in range(1, len(list(list(devices.columns)[1]))):
        device_ips.append(list(list(devices.columns)[1])[i].value)

    testbed = Create_Testbed(user,password,hostnames,device_ips,jump_bool,jump_ip)
    try:
        global tb
        tb = load(testbed)
    except Exception as e:
        logging.debug('Exception on load testbed:\n {}'.format(e))

    os.system('rm Output.xlsx')
    global sheet
    out_wb = Workbook()
    sheet = out_wb.active
    sheet.title = "Parsed"
    sheet.append(["Hostname","Underlay IP"])

    result = pool_connection(12,hostnames)
    if result:
        print("Success")
        out_wb.save(filename = "Output.xlsx")

if __name__ == "__main__":
    main()

